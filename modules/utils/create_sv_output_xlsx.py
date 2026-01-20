#!/usr/bin/env python3
"""create_sv_output_xlsx.py

Create an SV summary Excel workbook from three TSV files produced by
modules/sv_calling.nf (vcf_to_table).

This version follows the updated business logic:
- Split SVs into sheets by type (Insertions/Deletions/Replacements/Inversions/Translocations).
- Assign event IDs per sheet by sorting events by their primary start coordinate.
- Primary grouping is assembly-driven: for each assembly interval, collect all short/long
  calls of the same SV category that are fully contained in the assembly interval and
  compute medians/std.
- Remaining long calls (not assigned to any assembly interval) form events; short calls
  fully contained in those long calls are merged into the same event/row (same event_id).
- Remaining short calls form short-only events.

Template handling
-----------------
If --template is provided, the script:
- loads it,
- preserves rows 1-2 in each sheet,
- deletes rows 3+ (example data),
- writes new data starting at row 3 using the row-2 keys as the schema.

If --template is not provided, the script will create a workbook from scratch using a
schema matching the supplied template (including its keys).

Notes about upstream TSVs
-------------------------
Long TSV (updated): chrom, start, end, svtype, info_svtype, supporting_reads, score
Short TSV:          chrom, start, end, svtype, info_svtype, debreak_type, supporting_reads, score
SyRI TSV:           chrom, start, end, svtype

Supporting reads:
- If a value contains commas (e.g. "0,00,0793,1957"), the 3rd number (index 2) is used.

Lengths:
- Length is computed as inclusive: end - start + 1 (per review comments).

Translocations:
- SyRI "TRANS*" and "TRANSAL*" are treated as translocations.
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import statistics
from dataclasses import dataclass, field
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook


# Template-matching schema (row 2 keys) extracted from the provided output.xlsx template.
TEMPLATE_KEYS: Dict[str, List[str]] = {
    "Insertions": [
        "event_id",
        "event_type",
        "sequence_id",
        "asm_start",
        "asm_end",
        "asm_length_bp",
        "long_reads_start_median",
        "long_reads_start_std",
        "long_reads_end_median",
        "long_reads_end_std",
        "long_reads_length_median",
        "short_reads_start_median",
        "short_reads_start_std",
        "short_reads_end_median",
        "short_reads_end_std",
        "short_reads_length_median",
        "long_reads_support",
        "long_reads_score",
        "short_reads_support",
        "short_reads_score",
        "asm_copy_number_estimate",
        "long_reads_copy_number_estimate",
        "short_reads_copy_number_estimate",
    ],
    "Deletions": [
        "event_id",
        "event_type",
        "sequence_id",
        "asm_start",
        "asm_end",
        "asm_length_bp",
        # template typo preserved:
        "long_short_readss_start_median",
        "long_reads_start_std",
        "long_reads_end_median",
        "long_reads_end_std",
        "long_reads_length_median",
        "short_reads_start_median",
        "short_reads_start_std",
        "short_reads_end_median",
        "short_reads_end_std",
        "short_reads_length_median",
        "assembly_score",
        "long_reads_support",
        "long_reads_score",
        "short_reads_support",
        "short_reads_score",
        "asm_copy_number_estimate",
        "long_reads_copy_number_estimate",
        "short_reads_copy_number_estimate",
    ],
    "Replacements": [
        "event_id",
        "event_type",
        "sequence_id",
        "asm_del_start",
        "asm_del_end",
        "asm_del_length",
        "asm_ins_start",
        "asm_ins_end",
        "asm_ins_length",
        "long_reads_del_start_median",
        "long_reads_del_start_std",
        "long_reads_del_end_median",
        "long_reads_del_end_std",
        "long_reads_del_length_median",
        "long_reads_ins_start_median",
        "long_reads_ins_start_std",
        "long_reads_ins_end_median",
        "long_reads_ins_end_std",
        "long_reads_ins_length_median",
        "short_reads_del_start_median",
        "short_reads_del_start_std",
        "short_reads_del_end_median",
        "short_reads_del_end_std",
        "short_reads_del_length_median",
        "short_reads_ins_start_median",
        "short_reads_ins_start_std",
        "short_reads_ins_end_median",
        "short_reads_ins_end_std",
        "short_reads_ins_length_median",
        "assembly_support",
        "assembly_score",
        "long_reads_support",
        "long_reads_score",
        "short_reads_support",
        "short_reads_score",
        "asm_copy_number_estimate",
        "long_reads_copy_number_estimate",
        "short_reads_copy_number_estimate",
    ],
    "Inversions": [
        "event_id",
        "event_type",
        "sequence_id",
        "asm_start",
        "asm_end",
        "asm_length_bp",
        "long_reads_start_median",
        "long_reads_start_std",
        "long_reads_end_median",
        "long_reads_end_std",
        "long_reads_length_median",
        "short_reads_start_median",
        "short_reads_start_std",
        "short_reads_end_median",
        "short_reads_end_std",
        "short_reads_length_median",
        "assembly_support",
        "assembly_score",
        "long_reads_support",
        "long_reads_score",
        "short_reads_support",
        "short_reads_score",
        "asm_copy_number_estimate",
        "long_reads_copy_number_estimate",
        "short_reads_copy_number_estimate",
    ],
    "Translocations": [
        "event_id",
        "event_type",
        "origin_sequence_id",
        "destination_sequence_id",
        "asm_origin_start",
        "asm_origin_end",
        "asm_dest_start",
        "asm_dest_end",
        "asm_length",
        "long_reads_origin_start_median",
        "long_reads_origin_start_std",
        "long_reads_origin_end_median",
        "long_reads_origin_end_std",
        "long_reads_dest_start_median",
        "long_reads_dest_start_std",
        "long_reads_dest_end_median",
        "long_reads_dest_end_std",
        "long_reads_length_median",
        "short_reads_origin_start_median",
        "short_reads_origin_start_std",
        "short_reads_origin_end_median",
        "short_reads_origin_end_std",
        "short_reads_dest_start_median",
        "short_reads_dest_start_std",
        "short_reads_dest_end_median",
        "short_reads_dest_end_std",
        "short_reads_length_median",
        "assembly_support",
        "assembly_score",
        "long_reads_support",
        "long_reads_score",
        "short_reads_support",
        "short_reads_score",
        "asm_copy_number_estimate",
        "long_reads_copy_number_estimate",
        "short_reads_copy_number_estimate",
    ],
}


# Canonical SV categories used internally for grouping
CAT_INS = "INS"
CAT_DEL = "DEL"
CAT_INV = "INV"
CAT_REP = "REP"
CAT_TRL = "TRL"

CAT_TO_SHEET = {
    CAT_INS: "Insertions",
    CAT_DEL: "Deletions",
    CAT_INV: "Inversions",
    CAT_REP: "Replacements",
    CAT_TRL: "Translocations",
}

CAT_TO_EVENT_TYPE_WORD = {
    CAT_INS: "insertion",
    CAT_DEL: "deletion",
    CAT_INV: "inversion",
    CAT_REP: "replacement",
    CAT_TRL: "translocation",
}

CAT_TO_ID_PREFIX = {
    CAT_INS: "INS",
    CAT_DEL: "DEL",
    CAT_INV: "INV",
    CAT_REP: "REP",
    CAT_TRL: "TRL",  # matches the provided template
}


# Mapping for SyRI svtype labels (prefixes) to canonical categories
SYRI_PREFIX_TO_CAT = {
    "TRANSAL": CAT_TRL,
    "TRANS": CAT_TRL,
    "CPL": CAT_REP,  # discussed as duplication-like; fits existing Replacements sheet
    "CPG": CAT_REP,
    "DEL": CAT_DEL,
    "INV": CAT_INV,
    "INS": CAT_INS,
}

# SVTYPE tokens to canonical categories
TOKEN_TO_CAT = {
    "INS": CAT_INS,
    "DEL": CAT_DEL,
    "INV": CAT_INV,
    "DUP": CAT_REP,  # no DUP tab in template; place into Replacements
    "BND": CAT_TRL,
    "TRA": CAT_TRL,
    "TRANS": CAT_TRL,
    "REP": CAT_REP,
}


def _strip(x: Any) -> str:
    return "" if x is None else str(x).strip()


def to_int(x: Any) -> Optional[int]:
    s = _strip(x)
    if not s or s.lower() == "nan" or s == ".":
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def to_float(x: Any) -> Optional[float]:
    s = _strip(x)
    if not s or s.lower() == "nan" or s == ".":
        return None
    try:
        return float(s)
    except Exception:
        return None


def parse_supporting_reads(x: Any) -> Optional[int]:
    """
    supporting_reads can be:
    - an int-like string
    - or 4 comma-separated values; the 3rd is the supporting reads count.
    """
    s = _strip(x)
    if not s or s.lower() == "nan" or s == ".":
        return None
    if "," in s:
        parts = [p for p in s.split(",") if p != ""]
        if len(parts) >= 3:
            return to_int(parts[2])
        return to_int(parts[-1]) if parts else None
    return to_int(s)


def inclusive_length(start: Optional[int], end: Optional[int]) -> Optional[int]:
    if start is None or end is None:
        return None
    return (end - start) + 1


def normalize_category(
    svtype: Any,
    info_svtype: Any = None,
    debreak_type: Any = None,
    *,
    is_syri: bool = False,
) -> Optional[str]:
    """
    Normalize upstream svtype representations to a canonical category.
    Priority:
      1) info_svtype token if present
      2) debreak_type if it's <SVTYPE>
      3) token parsed from svtype (caller ids like cuteSV.INS.0, INV000001)
      4) SyRI prefix mapping (TRANS*, TRANSAL*, CPL*, CPG*, etc.)
    """
    info = _strip(info_svtype).upper()
    if info in TOKEN_TO_CAT:
        return TOKEN_TO_CAT[info]

    db = _strip(debreak_type).upper()
    m = re.match(r"^<([A-Z0-9_]+)>$", db)
    if m:
        tok = m.group(1)
        if tok in TOKEN_TO_CAT:
            return TOKEN_TO_CAT[tok]

    raw = _strip(svtype)
    up = raw.upper()

    # SyRI: map by prefixes
    if is_syri:
        for pref, cat in SYRI_PREFIX_TO_CAT.items():
            if up.startswith(pref):
                return cat
        return None

    # Caller IDs / IDs: search for known tokens
    # Examples:
    # - cuteSV.INS.0
    # - Sniffles2.DUP.945S0
    # - INV00000001
    token_match = re.search(r"(INS|DEL|INV|DUP|BND|TRA|TRANS|REP)", up)
    if token_match:
        tok = token_match.group(1)
        return TOKEN_TO_CAT.get(tok)

    # Some formats use '.' separators; try split fallback
    for part in re.split(r"[\\W_\\.]+", up):
        if part in TOKEN_TO_CAT:
            return TOKEN_TO_CAT[part]

    return None


def read_tsv(path: str) -> List[Dict[str, Any]]:
    with open(path, newline="") as fh:
        rd = csv.DictReader(fh, delimiter="\t")
        return [dict(r) for r in rd]


@dataclass
class Call:
    chrom: str
    start: int
    end: int
    cat: str
    support: Optional[int] = None
    score: Optional[float] = None
    source: str = ""  # "short" or "long"

    @property
    def length(self) -> Optional[int]:
        return inclusive_length(self.start, self.end)


@dataclass
class AssemblyInterval:
    chrom: str
    start: int
    end: int
    cat: str
    svtype_raw: str = ""

    @property
    def length(self) -> Optional[int]:
        return inclusive_length(self.start, self.end)


@dataclass
class Event:
    cat: str
    chrom: str
    start: int
    end: int
    # optional assembly anchor
    asm_start: Optional[int] = None
    asm_end: Optional[int] = None
    asm_length: Optional[int] = None
    # supporting calls
    long_calls: List[Call] = field(default_factory=list)
    short_calls: List[Call] = field(default_factory=list)
    # translocation specific (unknown in current TSVs)
    dest_chrom: Optional[str] = None
    dest_start: Optional[int] = None
    dest_end: Optional[int] = None

    def primary_sort_key(self) -> Tuple[str, int, int]:
        return (self.chrom, self.start, self.end)


def contained_in(inner_start: int, inner_end: int, outer_start: int, outer_end: int) -> bool:
    return inner_start >= outer_start and inner_end <= outer_end


def assign_calls_to_interval(
    interval: AssemblyInterval,
    calls: Iterable[Call],
) -> List[Call]:
    """Return calls fully contained in an assembly interval (default behavior)."""
    out: List[Call] = []
    for c in calls:
        if c.cat != interval.cat:
            continue
        if c.chrom != interval.chrom:
            continue
        if contained_in(c.start, c.end, interval.start, interval.end):
            out.append(c)
    return out


def median_std(values: List[Optional[float]]) -> Tuple[Optional[float], Optional[float]]:
    vals = [v for v in values if v is not None]
    if not vals:
        return None, None
    if len(vals) == 1:
        return float(vals[0]), 0.0
    med = float(statistics.median(vals))
    mean = sum(vals) / len(vals)
    var = sum((v - mean) ** 2 for v in vals) / (len(vals) - 1)
    return med, math.sqrt(var)


def build_calls(rows: List[Dict[str, Any]], source: str) -> List[Call]:
    out: List[Call] = []
    for r in rows:
        chrom = _strip(r.get("chrom"))
        start = to_int(r.get("start"))
        end = to_int(r.get("end"))
        if not chrom or start is None or end is None:
            continue

        cat = normalize_category(
            r.get("svtype"),
            r.get("info_svtype"),
            r.get("debreak_type"),
            is_syri=False,
        )
        if not cat:
            continue

        sup = parse_supporting_reads(r.get("supporting_reads"))
        score = to_float(r.get("score"))
        out.append(Call(chrom=chrom, start=start, end=end, cat=cat, support=sup, score=score, source=source))
    return out


def build_assembly(rows: List[Dict[str, Any]]) -> List[AssemblyInterval]:
    out: List[AssemblyInterval] = []
    for r in rows:
        chrom = _strip(r.get("chrom"))
        start = to_int(r.get("start"))
        end = to_int(r.get("end"))
        if not chrom or start is None or end is None:
            continue
        raw = _strip(r.get("svtype"))
        cat = normalize_category(raw, is_syri=True)
        if not cat:
            continue
        out.append(AssemblyInterval(chrom=chrom, start=start, end=end, cat=cat, svtype_raw=raw))
    # Sort to ensure stable event ID assignment
    out.sort(key=lambda x: (x.chrom, x.start, x.end))
    return out


def create_events(
    assembly: List[AssemblyInterval],
    long_calls: List[Call],
    short_calls: List[Call],
) -> List[Event]:
    """
    Create events using:
    1) Assembly-driven intervals first
    2) Remaining long calls
    3) Remaining short calls
    """
    remaining_long = list(long_calls)
    remaining_short = list(short_calls)

    events: List[Event] = []

    # 1) Assembly-driven events
    for iv in assembly:
        in_long = assign_calls_to_interval(iv, remaining_long)
        in_short = assign_calls_to_interval(iv, remaining_short)

        if in_long:
            long_set = set(id(c) for c in in_long)
            remaining_long = [c for c in remaining_long if id(c) not in long_set]
        if in_short:
            short_set = set(id(c) for c in in_short)
            remaining_short = [c for c in remaining_short if id(c) not in short_set]

        # Create event even if no supporting calls (important for translocations in SyRI)
        ev = Event(
            cat=iv.cat,
            chrom=iv.chrom,
            start=iv.start,
            end=iv.end,
            asm_start=iv.start,
            asm_end=iv.end,
            asm_length=iv.length,
            long_calls=in_long,
            short_calls=in_short,
        )
        events.append(ev)

    # 2) Remaining long calls -> events (each long call defines an event)
    remaining_long.sort(key=lambda c: (c.cat, c.chrom, c.start, c.end))
    used_short_ids: set[int] = set()

    for lc in remaining_long:
        # absorb short calls fully contained in long call of same cat/chrom
        absorbed: List[Call] = []
        for sc in remaining_short:
            if id(sc) in used_short_ids:
                continue
            if sc.cat != lc.cat:
                continue
            if sc.chrom != lc.chrom:
                continue
            if contained_in(sc.start, sc.end, lc.start, lc.end):
                absorbed.append(sc)
                used_short_ids.add(id(sc))

        ev = Event(
            cat=lc.cat,
            chrom=lc.chrom,
            start=lc.start,
            end=lc.end,
            asm_start=None,
            asm_end=None,
            asm_length=None,
            long_calls=[lc],
            short_calls=absorbed,
        )
        events.append(ev)

    remaining_short = [sc for sc in remaining_short if id(sc) not in used_short_ids]

    # 3) Remaining short calls -> short-only events
    remaining_short.sort(key=lambda c: (c.cat, c.chrom, c.start, c.end))
    for sc in remaining_short:
        ev = Event(
            cat=sc.cat,
            chrom=sc.chrom,
            start=sc.start,
            end=sc.end,
            asm_start=None,
            asm_end=None,
            asm_length=None,
            long_calls=[],
            short_calls=[sc],
        )
        events.append(ev)

    return events


def load_or_create_workbook(template: Optional[str]) -> Workbook:
    if template:
        return load_workbook(template)

    wb = Workbook()
    # remove default sheet
    if wb.active:
        wb.remove(wb.active)

    for sheet_name, keys in TEMPLATE_KEYS.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["" for _ in keys])  # row 1 descriptions blank
        ws.append(keys)  # row 2 schema keys
    return wb


def clear_data_rows(ws) -> None:
    # keep rows 1-2, delete rows 3+
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)


def sheet_keys(ws) -> List[str]:
    # Prefer the template's row-2 keys. Fall back to embedded.
    keys: List[str] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v is None:
            continue
        keys.append(str(v))
    if keys:
        return keys
    return TEMPLATE_KEYS.get(ws.title, [])


def event_to_row_dict(ev: Event) -> Dict[str, Any]:
    """
    Convert an Event into a dict keyed by template schema keys.
    Unknown fields are left as None.
    """
    cat = ev.cat
    sheet = CAT_TO_SHEET[cat]
    event_type_word = CAT_TO_EVENT_TYPE_WORD[cat]

    # Assembly coordinates: if event is assembly-derived use asm_*, else fall back to event bounds
    asm_start = ev.asm_start if ev.asm_start is not None else ev.start
    asm_end = ev.asm_end if ev.asm_end is not None else ev.end
    asm_len = ev.asm_length if ev.asm_length is not None else inclusive_length(asm_start, asm_end)

    # Build call stats
    long_starts = [c.start for c in ev.long_calls]
    long_ends = [c.end for c in ev.long_calls]
    long_lens = [inclusive_length(c.start, c.end) for c in ev.long_calls]
    long_sups = [c.support for c in ev.long_calls]
    long_scores = [c.score for c in ev.long_calls]

    short_starts = [c.start for c in ev.short_calls]
    short_ends = [c.end for c in ev.short_calls]
    short_lens = [inclusive_length(c.start, c.end) for c in ev.short_calls]
    short_sups = [c.support for c in ev.short_calls]
    short_scores = [c.score for c in ev.short_calls]

    lr_start_med, lr_start_std = median_std([float(v) for v in long_starts] if long_starts else [])
    lr_end_med, lr_end_std = median_std([float(v) for v in long_ends] if long_ends else [])
    lr_len_med, _ = median_std([float(v) for v in long_lens if v is not None])
    lr_sup_med, _ = median_std([float(v) for v in long_sups if v is not None])
    lr_score_med, _ = median_std([float(v) for v in long_scores if v is not None])

    sr_start_med, sr_start_std = median_std([float(v) for v in short_starts] if short_starts else [])
    sr_end_med, sr_end_std = median_std([float(v) for v in short_ends] if short_ends else [])
    sr_len_med, _ = median_std([float(v) for v in short_lens if v is not None])
    sr_sup_med, _ = median_std([float(v) for v in short_sups if v is not None])
    sr_score_med, _ = median_std([float(v) for v in short_scores if v is not None])

    row: Dict[str, Any] = {}

    if sheet in ("Insertions", "Inversions"):
        row.update(
            {
                "event_type": event_type_word,
                "sequence_id": ev.chrom,
                "asm_start": asm_start,
                "asm_end": asm_end,
                "asm_length_bp": asm_len,
                "long_reads_start_median": lr_start_med,
                "long_reads_start_std": lr_start_std,
                "long_reads_end_median": lr_end_med,
                "long_reads_end_std": lr_end_std,
                "long_reads_length_median": lr_len_med,
                "short_reads_start_median": sr_start_med,
                "short_reads_start_std": sr_start_std,
                "short_reads_end_median": sr_end_med,
                "short_reads_end_std": sr_end_std,
                "short_reads_length_median": sr_len_med,
                "long_reads_support": lr_sup_med,
                "long_reads_score": lr_score_med,
                "short_reads_support": sr_sup_med,
                "short_reads_score": sr_score_med,
            }
        )

    elif sheet == "Deletions":
        # template has a key typo for long start median
        row.update(
            {
                "event_type": event_type_word,
                "sequence_id": ev.chrom,
                "asm_start": asm_start,
                "asm_end": asm_end,
                "asm_length_bp": asm_len,
                "long_short_readss_start_median": lr_start_med,
                "long_reads_start_std": lr_start_std,
                "long_reads_end_median": lr_end_med,
                "long_reads_end_std": lr_end_std,
                "long_reads_length_median": lr_len_med,
                "short_reads_start_median": sr_start_med,
                "short_reads_start_std": sr_start_std,
                "short_reads_end_median": sr_end_med,
                "short_reads_end_std": sr_end_std,
                "short_reads_length_median": sr_len_med,
                "assembly_score": None,
                "long_reads_support": lr_sup_med,
                "long_reads_score": lr_score_med,
                "short_reads_support": sr_sup_med,
                "short_reads_score": sr_score_med,
            }
        )

    elif sheet == "Replacements":
        # With current upstream TSVs we don't have separate del/ins intervals, so we treat asm interval as the del part
        row.update(
            {
                "event_type": event_type_word,
                "sequence_id": ev.chrom,
                "asm_del_start": asm_start,
                "asm_del_end": asm_end,
                "asm_del_length": asm_len,
                "asm_ins_start": None,
                "asm_ins_end": None,
                "asm_ins_length": None,
                "long_reads_del_start_median": lr_start_med,
                "long_reads_del_start_std": lr_start_std,
                "long_reads_del_end_median": lr_end_med,
                "long_reads_del_end_std": lr_end_std,
                "long_reads_del_length_median": lr_len_med,
                "long_reads_ins_start_median": None,
                "long_reads_ins_start_std": None,
                "long_reads_ins_end_median": None,
                "long_reads_ins_end_std": None,
                "long_reads_ins_length_median": None,
                "short_reads_del_start_median": sr_start_med,
                "short_reads_del_start_std": sr_start_std,
                "short_reads_del_end_median": sr_end_med,
                "short_reads_del_end_std": sr_end_std,
                "short_reads_del_length_median": sr_len_med,
                "short_reads_ins_start_median": None,
                "short_reads_ins_start_std": None,
                "short_reads_ins_end_median": None,
                "short_reads_ins_end_std": None,
                "short_reads_ins_length_median": None,
                "assembly_support": None,
                "assembly_score": None,
                "long_reads_support": lr_sup_med,
                "long_reads_score": lr_score_med,
                "short_reads_support": sr_sup_med,
                "short_reads_score": sr_score_med,
            }
        )

    elif sheet == "Translocations":
        row.update(
            {
                "event_type": event_type_word,
                "origin_sequence_id": ev.chrom,
                "destination_sequence_id": ev.dest_chrom,
                "asm_origin_start": asm_start,
                "asm_origin_end": asm_end,
                "asm_dest_start": ev.dest_start,
                "asm_dest_end": ev.dest_end,
                "asm_length": asm_len,
                "long_reads_origin_start_median": lr_start_med,
                "long_reads_origin_start_std": lr_start_std,
                "long_reads_origin_end_median": lr_end_med,
                "long_reads_origin_end_std": lr_end_std,
                "long_reads_dest_start_median": None,
                "long_reads_dest_start_std": None,
                "long_reads_dest_end_median": None,
                "long_reads_dest_end_std": None,
                "long_reads_length_median": lr_len_med,
                "short_reads_origin_start_median": sr_start_med,
                "short_reads_origin_start_std": sr_start_std,
                "short_reads_origin_end_median": sr_end_med,
                "short_reads_origin_end_std": sr_end_std,
                "short_reads_dest_start_median": None,
                "short_reads_dest_start_std": None,
                "short_reads_dest_end_median": None,
                "short_reads_dest_end_std": None,
                "short_reads_length_median": sr_len_med,
                "assembly_support": None,
                "assembly_score": None,
                "long_reads_support": lr_sup_med,
                "long_reads_score": lr_score_med,
                "short_reads_support": sr_sup_med,
                "short_reads_score": sr_score_med,
            }
        )

    return row


def write_events_to_workbook(wb: Workbook, events: List[Event]) -> None:
    # Organize events by sheet and assign event IDs based on sorted occurrence
    events_by_sheet: Dict[str, List[Event]] = {name: [] for name in TEMPLATE_KEYS.keys()}
    for ev in events:
        sheet = CAT_TO_SHEET.get(ev.cat)
        if sheet:
            events_by_sheet[sheet].append(ev)

    # Clear and write each sheet
    for sheet_name, evs in events_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        keys = sheet_keys(ws)
        clear_data_rows(ws)

        # Sort events and assign IDs
        evs.sort(key=lambda e: e.primary_sort_key())
        prefix = CAT_TO_ID_PREFIX[[k for k, v in CAT_TO_SHEET.items() if v == sheet_name][0]]

        for i, ev in enumerate(evs, 1):
            row_dict = event_to_row_dict(ev)
            row_dict["event_id"] = f"{prefix}_{i}"
            # Ensure row_dict contains any template-specific alias keys if needed
            # (e.g. no-op here except deletions already uses the typo key)
            ws.append([row_dict.get(k) for k in keys])


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--assembly", required=True, help="SyRI TSV (ref_x_modsyri_sv_summary.tsv)")
    ap.add_argument("--short", required=True, help="Short-read TSV (*_sv_short_read_sv_summary.tsv)")
    ap.add_argument("--long", required=True, help="Long-read TSV (*_sv_long_read_sv_summary.tsv)")
    ap.add_argument("--out", required=True, help="Output XLSX path")
    ap.add_argument("--template", default=None, help="Optional template XLSX (output.xlsx)")

    args = ap.parse_args()

    assembly_rows = read_tsv(args.assembly)
    short_rows = read_tsv(args.short)
    long_rows = read_tsv(args.long)

    assembly = build_assembly(assembly_rows)
    short_calls = build_calls(short_rows, source="short")
    long_calls = build_calls(long_rows, source="long")

    events = create_events(assembly, long_calls, short_calls)

    wb = load_or_create_workbook(args.template)
    write_events_to_workbook(wb, events)
    wb.save(args.out)


if __name__ == "__main__":
    main()
